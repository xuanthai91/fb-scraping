import os
import datetime
import shutil
import time
import glob
import json
import csv
from time import sleep
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
import selenium.webdriver.support.expected_conditions as EC
import pickle
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request

#base project info
base_path = os.getcwd()
print(base_path)
fmaster = os.path.join(base_path, '.\\Master.xlsx')
dt_master = load_workbook(fmaster,data_only=True)
wb_sheet = dt_master['Master']
driver = webdriver.Chrome(executable_path = base_path + '.\\function\\chromedriver.exe')

#create folder
str_today = datetime.datetime.now().strftime('%Y%m%d')
sub_folder = ['raw','output','tmp']
main_folder = os.path.abspath(str_today)
tmp_folder = os.path.abspath(str_today + '\\tmp')
raw_folder = os.path.abspath(str_today + '\\raw')
output_folder = os.path.abspath(str_today + '\\output')
for i in sub_folder:
    os.makedirs(main_folder + '\\' + i, exist_ok=True)
print('Create folder %s OK!!!' %str_today)

#setting chrome
def get_chrome(dl_folder = None, load_cookies = False):
    chrome_opt = webdriver.ChromeOptions()
    chrome_opt.add_argument('--disable-infobars')
    chrome_opt.add_argument('--start-maximized')

    if dl_folder is not None:
        prefs = {'download.default_directory': dl_folder}
        chrome_opt.add_experimental_option('prefs', prefs)
    
    if load_cookies:
        pass
    else:
        chrome_opt.add_argument('--incognito')
    return webdriver.Chrome(executable_path= base_path + '.\\function\\chromedriver.exe',
                            options=chrome_opt)

#convert time in setting file
def time_convert(str_time: str):
    if str_time.lower().__contains__('custom'):
        end_day = datetime.datetime.strptime(str_time.split(';')[-1], '%Y%m%d')
        start_day = datetime.datetime.strptime(str_time.split(';')[1], '%Y%m%d')
    else:
        print('Format time %s was wrong' %str_time)
    return[datetime.datetime.strftime(start_day, '%Y/%m/%d'), datetime.datetime.strftime(end_day, '%Y/%m/%d')]

#read setting and get info
def read_settei():
    acc_info = []
    for i in range(2, wb_sheet.max_row + 1):
        sub_info = []
        if wb_sheet.cell(row=i, column=1).value == 'ON':
            prj_name = wb_sheet.cell(row=i, column=3).value
            sub_info.append(prj_name)

            kikan = time_convert(wb_sheet.cell(row=i, column=4).value)
            sub_info.append(kikan[0])
            sub_info.append(kikan[1])

            acc_id = wb_sheet.cell(row=i, column=5).value
            sub_info.append(acc_id)

            business_id = wb_sheet.cell(row=i, column=6).value
            sub_info.append(business_id)

            cl_preset = wb_sheet.cell(row=i, column=7).value
            sub_info.append(cl_preset)

            tab_info = wb_sheet.cell(row=i, column=8).value
            sub_info.append(tab_info)

            search_key = wb_sheet.cell(row=i, column=9).value
            if search_key is None:
                sub_info.append('None')
            else:
                sub_info.append(search_key)

            spreadsheet_id = wb_sheet.cell(row=i, column=10).value
            sub_info.append(spreadsheet_id)

            fname = wb_sheet.cell(row=i, column=11).value
            sub_info.append(fname)
            acc_info.append(sub_info)
    return acc_info

#print(read_settei())

def clean_folder(input_folder):
    for file in os.listdir(input_folder):
        os.unlink(os.path.join(input_folder, file))
    return print('Clean', input_folder, 'done!')

def move_and_rename(dlpath, des_folder, newfilename, ext):
    get_file = False
    retry = 0
    while get_file == False:
        try:
            currentfolder = glob.glob(dlpath + '\\*')
            des_file = des_folder + newfilename + ext
            if newfilename == '':
                newfilename = currentfolder[0].replace('\\', '/').split('/')[-1]
                des_file = des_folder + newfilename
            if len(currentfolder) == 0 or ('.crdownload' in currentfolder[0]):
                print('File downloading')
                retry += 1
                sleep(3)
            elif retry == 30:
                print('Download error: ', newfilename)
                return False
            else:
                get_file = True
        except:
            print("Can't get file(download failed)")
        
    try:
        currentfolder = currentfolder[0].replace('\\', '/')
        shutil.move(currentfolder, des_file)
        return des_file
    except:
        if '/' in newfilename:
            return False
        sleep(5)

def highlight(element):
    """Highlights (blinks) a Selenium Webdriver element"""

    browser = element._parent

    def appy_style(s):
        browser.execute_script("arguments[0].setAttribute('style', arguments[1]);", element, s)

    original_style = element.get_attribute('style')
    appy_style("background: yellow; border: 2px solid red;")
    sleep(.2)
    appy_style(original_style)

def download_fb(sheet_master):
    driver = get_chrome(tmp_folder, False)
    driver.get('https://www.google.com/')
    cookies = json.load(open(base_path + '.\\function\\facebook_cookies.txt', 'r'))
    for c in cookies:
        try:
            del c['expiry']
        except:
            pass
        try:
            del c['sameSite']
        except:
            pass
        driver.add_cookie(c)
    sleep(1)
    
    a = 0
    for i in sheet_master:
        if i[6] == 'reporting':
            if i[7] == 'None':
                link = 'https://business.facebook.com/adsmanager/reporting/view?act={}&business_id={}&selected_report_id={}&time_range={}_{}'.format(i[3], i[4], i[5],
                    i[1].replace("/", "-"), (datetime.datetime.strptime(i[2], "%Y/%m/%d") + datetime.timedelta(days=1)).strftime("%Y/%m/%d").replace("/", "-"))
            else:
                link = 'https://business.facebook.com/adsmanager/reporting/view?act={}&business_id={}&selected_report_id={}&time_range={}_{}&filter_set=SEARCH_BY_CAMPAIGN_GROUP_NAME-STRING%1ECONTAIN%1E"'"{}"'"'.format(
                    i[3], i[4], i[5], i[1].replace("/", "-"), (datetime.datetime.strptime(i[2], "%Y/%m/%d") + datetime.timedelta(days=1)).strftime("%Y/%m/%d").replace("/", "-"), i[7].replace('None', ''))
        else:
            if i[7] == 'None':
                link = 'https://business.facebook.com/adsmanager/manage/{}?act={}&business_id={}&column_preset={}&date={}_{}'.format(i[6], i[3], i[4], i[5],
                    i[1].replace("/", "-"), (datetime.datetime.strptime(i[2], "%Y/%m/%d") + datetime.timedelta(days=1)).strftime("%Y/%m/%d").replace("/", "-"))
            else:
                link = 'https://business.facebook.com/adsmanager/manage/{}?act={}&business_id={}&column_preset={}&date={}_{}&filter_set=SEARCH_BY_CAMPAIGN_GROUP_NAME-STRING%1ECONTAIN%1E"'"{}"'"'.format(i[6], 
                    i[3], i[4], i[5], i[1].replace("/", "-"), (datetime.datetime.strptime(i[2], "%Y/%m/%d") + datetime.timedelta(days=1)).strftime("%Y/%m/%d").replace("/", "-"), i[7].replace('None', ''))
        driver.get(link)
        a += 1
        print('Downloading---', a, '/', len(sheet_master))
        
        for l in {1, 2, 3}:
            count = 0
            while True:
                count += 1
                sleep(1)
                if driver.execute_script('return document.readyState') == 'complete' or count > 30:
                    break
        sleep(3)
        
        export_button = "//div[text()='レポート']"
        
        export_confirm = "//button[@data-testid='export-confirm-button'] | " \
                            "//div[@role='dialog']//div[text()='エクスポート']/ancestor::button | " \
                            "//span[text()='エクスポート']/ancestor::button"

        retry = 0
        while True:
            elem = WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, export_button)))
            highlight(elem)
            elem.click()
            sleep(3)
            try:
                test_button = driver.find_elements_by_xpath("//div[text()='テーブルデータをエクスポート...']")
            except:
                test_button =[]

            retry += 1
            if len(test_button)> 0 or retry == 5:
                break
            
        sleep(1)
        
        if i[6] != 'reporting':
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, '//div[text()="テーブルデータをエクスポート..."]'))).click()
            sleep(1)
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, '//input[@value = "csv"]'))).click()
        sleep(1)
        
        if i[6] == 'reporting':
            driver.find_element_by_xpath("//input[@value='csv']//ancestor::div[@role='dialog']//*[text()='キャンセル']//ancestor::button").send_keys(Keys.TAB + Keys.ENTER)
        else:
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, export_confirm))).click()

        try:
            WebDriverWait(driver, 120).until(EC.invisibility_of_element_located((By.XPATH, "//div[@name='progress']")))
        except:
            pass
        sleep(3)
        
        move_and_rename(tmp_folder, raw_folder + '\\', i[9], '.csv')

    return print('Download OK!')

def data_prcess():
    header = ['date', 'company', 'cost']
    newFile = output_folder + '.\\out_put.csv'
    today = datetime.datetime.now().strftime('%Y/%m/%d')
    with open(newFile, 'a', newline = '', ) as nf:
        writeFile = csv.writer(nf)
        writeFile.writerow(header)
    for i in read_settei():
        company_name = i[9].split('_')[0]
        myFile = raw_folder + '\\' + i[9] + '.csv'
        print(myFile)
        with open(myFile, newline='', encoding='utf-8') as f:
            read_data = csv.reader(f)
            read_line = []
            for row in read_data:
                read_line.append(row)
            position = read_line[0].index('消化金額 (JPY)')
            
            sub_data = []
            for i in read_line:
                sub_data.append(i[position])
            
            sums = 0
            for s in sub_data[1:]:
                sums += int(s)
            print(sums)

        with open(newFile, 'a', newline = '', ) as nf:
            writeFile = csv.writer(nf)
            writeFile.writerow([today, company_name, sums])
    return newFile
    #print('Data Processed Done!')

def call_gsuite():
    # If modifying these scopes, delete the file token.pickle.
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

    creds = None
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first time.
    if os.path.exists('.\\token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)

    service = build('sheets', 'v4', credentials=creds)
    
    return service

    # Call the Sheets API
    #sheet = service.spreadsheets()
    #result = sheet.values().get(spreadsheetId=SPREADSHEET_ID,
                                #range=RANGE_NAME).execute()
    #values = result.get('values', [])

class gsuite_service(object):
    def __init__(self, service):
        self.service = service
    
    def addData_to_gsheet(self, sheet_id, sheet_range, array):
        body = {
            "majorDimension": "ROWS",
            "values": array
        }
        result = self.service.spreadsheets().values().update(
                spreadsheetId=sheet_id, range=sheet_range,
                valueInputOption='USER_ENTERED', body=body).execute()
        print('{0} cells updated.'.format(result.get('updatedCells')))
        return result
    
    def check_gsheet_exist(self, sheet_id, sheet_range):
        self.service.spreadsheets().get(spreadsheetId=sheet_id, ranges=sheet_range).execute()
    
    def clear_gsheet(self, sheet_id, sheet_range):
        self.service.spreadsheets().values().clear(spreadsheetId=sheet_id, range=sheet_range, body={}).execute()

if __name__ == "__main__":
    clean_folder(raw_folder)
    clean_folder(output_folder)
    download_fb(read_settei())
    data_prcess()
    service = gsuite_service(call_gsuite())
    output_file = output_folder + '.\\out_put.csv'
    
    with open(output_file, newline='', encoding='utf-8') as file:
        read_file = csv.reader(file)
        _data = []
        for row in read_file:
            _data.append(row)
        print(_data)

    for i in read_settei():
        _spreadsheetID = i[8]
        service.clear_gsheet(_spreadsheetID, 'Data!A1:E')
        service.addData_to_gsheet(_spreadsheetID, 'Data!A1:E', _data)